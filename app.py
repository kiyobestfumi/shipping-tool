import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import json
import base64
import requests
import datetime

# --- 【最新版】場所マスターデータの定義 ---
LOCATION_MASTER = {
    "マスターを使用しない (手入力用)": {"name": "", "address": "", "tel": "", "contact": ""},
    "エースロジコム 横浜": {
        "name": "株式会社エースロジコム 横浜事業所", 
        "address": "横浜市鶴見区大黒埠頭22番YCC物流棟109号",
        "tel": "045-510-2158", 
        "contact": "橋爪・定方"
    },
    "鈴江コーポレーション 横浜": {
        "name": "鈴江コーポレーション 新杉田埠頭倉庫営業所", 
        "address": "横浜市金沢区鳥浜町11番地",
        "tel": "045-774-1371", 
        "contact": ""
    },
}

# --- 【新規追加】搬入先CFSマスターデータの定義 ---
CFS_MASTER = {
    "エースロジコム東京": """＜搬入先CFS＞ 
株式会社エースロジコム東京倉庫(1FWT5)
東京都大田区東海1-3-6プロロジスパーク東京大田内1階
TEL:03-5755-7419 / FAX:03-5755-7423
※貨物送り状には、ブッキングナンバーを明記し「ベストシッピング扱い」とご記載ください。""",
    
    "エースロジコム横浜": """＜搬入先CFS＞ 
株式会社エースロジコム横浜倉庫(2HI47)
神奈川県横浜市鶴見区大黒埠頭22番YCC物流棟109号
TEL:045-510-2158 / FAX:045-510-2159
※貨物送り状には、ブッキングナンバーを明記し「ベストシッピング扱い」とご記載ください。"""
}

# 港名から「, 国名」をカットする便利関数
def format_port_name(port_name):
    if not port_name:
        return ""
    # カンマで分割して、最初の部分（港名）だけを取り出す
    return port_name.split(",")[0].strip()

# 共通のAIデータ抽出関数
def call_gemini_api(api_key, uploaded_file):
    try:
        base64_image = base64.b64encode(uploaded_file.getvalue()).decode('utf-8')
        mime_type = uploaded_file.type
        
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        headers = {'Content-Type': 'application/json'}
        
        prompt_text = """添付の画像（物流手配書・B/L等）から以下の情報を抽出し、JSON形式でのみ出力してください。
※重要：```json などの説明文は一切含めず、必ず { から始まり } で終わる純粋なJSONテキストのみを出力してください。
情報が見当たらない場合は空文字("")にしてください。
{
  "vessel": "船名",
  "voy": "VOY No",
  "bl_no": "B/L NO または BOOKING NO",
  "etd": "ETD（出港日、またはスケジュール上の出発日）",
  "eta": "ETA（入港日、またはスケジュール上の到着日）",
  "shipper": "荷主名",
  "quantity": "外装個数（※超重要：画像内の該当項目において、【左側の枠内にある数字】が個数であり、【右側の枠内にある文字（PLTやCSなど）】が形状・単位です。必ずこの2つだけを抽出・合体させて「〇〇 PLT」のように出力してください。それ以外の数字は絶対に無視してください）",
  "weight": "重量 (数字のみ)",
  "volume": "体積 M3 (数字のみ)",
  "internal_memo": "社内メモの内容（あれば）",
  "place_of_receipt": "PLACE OF RECEIPTの場所",
  "cfs_cut": "CFS CUTの日付",
  "place_of_loading": "PLACE OF LOADINGの場所",
  "place_of_delivery": "PLACE OF DELIVERYの場所"
}"""

        payload = {
            "contents": [{
                "parts": [
                    {"text": prompt_text},
                    {"inline_data": {"mime_type": mime_type, "data": base64_image}}
                ]
            }]
        }
        
        response = requests.post(url, headers=headers, json=payload)
        response_data = response.json()
        
        if "error" in response_data:
            st.error(f"Googleサーバーからのエラー: {response_data['error']['message']}")
            return None
            
        result_text = response_data['candidates'][0]['content']['parts'][0]['text'].strip()
        
        if result_text.startswith("```json"):
            result_text = result_text.replace("```json", "")
        if result_text.endswith("```"):
            result_text = result_text.replace("```", "")
        result_text = result_text.strip()
        
        return json.loads(result_text)
    except Exception as e:
        st.error(f"エラーが発生しました: {e}")
        return None

# 画面設定
st.set_page_config(page_title="物流スクショ自動入力ツール", layout="centered")
st.title("🚢 物流スクショ ➡️ 業務自動化ツール")
st.write("集荷依頼書のスクショから、Excelファイルの作成や顧客連絡用テキストの生成を行います。")

# 1. APIキー入力
api_key = st.text_input("Gemini API Keyを入力してください", type="password")

# 2. 画像アップロード
st.header("📤 画像アップロード")
uploaded_file = st.file_uploader("スクショ画像をアップロードしてください", type=["jpg", "jpeg", "png"])

# 3. 目的の選択（画面切り替え用のスイッチ）
st.header("🔄 処理内容の選択")
mode = st.selectbox("どちらの機能を利用しますか？", ("集荷依頼書 (Excel作成)", "顧客連絡用テキスト作成"))

# --- 画面切り替え：集荷依頼書（Excel作成）モード ---
if mode == "集荷依頼書 (Excel作成)":
    st.header("📄 集荷依頼書") 
    
    option = st.selectbox("手配車輌", ("OLT", "混載便", "チャーター便"))

    col1, col2 = st.columns(2)
    with col1:
        pickup_date_input = st.date_input("集荷日", datetime.date.today())
    with col2:
        delivery_date_input = st.date_input("配達日", datetime.date.today() + datetime.timedelta(days=1))

    st.subheader("🏢 取引先・場所マスター選択")
    col3, col4 = st.columns(2)
    with col3:
        pickup_loc_key = st.selectbox("集荷場所（マスターから選択）", list(LOCATION_MASTER.keys()))
    with col4:
        delivery_loc_key = st.selectbox("配達場所（マスターから選択）", list(LOCATION_MASTER.keys()))

    if uploaded_file and api_key:
        if st.button("Excelを自動生成する"):
            with st.spinner("最新のGemini AIと通信してデータを抽出しています...（約10秒）"):
                data = call_gemini_api(api_key, uploaded_file)
                if data:
                    st.success("データの抽出が完了しました！")
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "集荷依頼書"

                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 1
                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                    
                    ws.page_margins.left = 0.4
                    ws.page_margins.right = 0.4
                    ws.page_margins.top = 0.6
                    ws.page_margins.bottom = 0.6

                    font_title = Font(name='ＭＳ Ｐゴシック', size=18, bold=True, underline='single')
                    font_normal = Font(name='ＭＳ Ｐゴシック', size=11)
                    font_bold = Font(name='ＭＳ Ｐゴシック', size=11, bold=True)
                    font_large = Font(name='ＭＳ Ｐゴシック', size=12)
                    
                    align_left = Alignment(horizontal='left', vertical='center')
                    align_right = Alignment(horizontal='right', vertical='center')
                    border_bottom = Border(bottom=Side(style='thin', color='000000'))

                    cols_width = {'A': 25, 'B': 2, 'C': 18, 'D': 10, 'E': 16, 'F': 2, 'G': 8, 'H': 25, 'I': 12, 'J': 18, 'K': 5}
                    for col, width in cols_width.items():
                        ws.column_dimensions[col].width = width

                    for row in range(1, 55):
                        for col in range(1, 12):
                            ws.cell(row=row, column=col).font = font_normal

                    ws.merge_cells('A1:J1')
                    ws["A1"] = "集荷依頼書"
                    ws["A1"].font = font_title
                    ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
                    ws.row_dimensions[1].height = 30

                    ws["A3"] = "株式会社エースロジコム　横浜事業所御中"
                    ws["A3"].font = font_bold
                    ws["A5"] = "      TEL: 045-510-2158"
                    ws["A6"] = "      FAX: 045-510-2159"
                    ws["H7"] = "株式会社ベストシッピング"
                    ws["H7"].font = font_bold
                    ws["A8"] = "いつもお世話になっております。"
                    ws["H8"] = "担当："
                    ws["A9"] = "下記の件、配送手配お願いいたします。"
                    ws["H10"] = "TEL: 03-5439-3710"
                    ws["H11"] = "FAX: 03-5439-3704"

                    labels = {
                        "A13": "船名", "I13": "VOY.", "A16": "B/L NO/BOOKING NO", "I16": "ETA/ETD", "A19": "荷主",
                        "A22": "M3･重量", "G22": "KGS", "J22": "M3", "A25": "手配車輌", "A28": "集荷日", "A31": "配達日",
                        "A34": "集荷場所", "A35": "住所", "D36": "TEL", "H36": "担当：",
                        "A39": "配達場所", "A40": "住所", "D41": "TEL", "H41": "担当：", "A44": "備　考"
                    }
                    for cell, text in labels.items():
                        ws[cell] = text
                        ws[cell].font = font_bold

                    def set_data(cell_start, cell_end, value, align=align_left, is_wrap=False):
                        ws.merge_cells(f"{cell_start}:{cell_end}")
                        ws[cell_start] = value
                        ws[cell_start].font = font_large
                        ws[cell_start].alignment = Alignment(horizontal=align.horizontal, vertical=align.vertical, wrap_text=is_wrap)
                        for col_idx in range(openpyxl.utils.column_index_from_string(cell_start[0]), openpyxl.utils.column_index_from_string(cell_end[0]) + 1):
                            ws.cell(row=int(cell_start[1:]), column=col_idx).border = border_bottom

                    set_data("C13", "H13", data.get("vessel", ""))
                    set_data("J13", "K13", data.get("voy", ""))
                    set_data("C16", "H16", data.get("bl_no", ""))
                    
                    etd_eta_combined = f"ETD: {data.get('etd', '')}\nETA: {data.get('eta', '')}"
                    set_data("J16", "K16", etd_eta_combined, align=align_left, is_wrap=True)
                    ws.row_dimensions[16].height = 32
                    
                    set_data("C19", "K19", data.get("shipper", ""))
                    set_data("C22", "D22", data.get("quantity", ""), align_right)
                    set_data("E22", "F22", data.get("weight", ""), align_right)
                    set_data("H22", "I22", data.get("volume", ""), align_right)
                    set_data("C25", "K25", option)
                    
                    set_data("C28", "K28", pickup_date_input.strftime('%Y/%m/%d'))
                    set_data("C31", "K31", delivery_date_input.strftime('%Y/%m/%d'))
                    
                    pickup_info = LOCATION_MASTER[pickup_loc_key]
                    set_data("C34", "K34", pickup_info["name"]) 
                    set_data("C35", "K35", pickup_info["address"]) 
                    set_data("E36", "F36", pickup_info["tel"]) 
                    set_data("I36", "K36", pickup_info["contact"]) 
                    
                    delivery_info = LOCATION_MASTER[delivery_loc_key]
                    set_data("C39", "K39", delivery_info["name"]) 
                    set_data("C40", "K40", delivery_info["address"]) 
                    set_data("E41", "F41", delivery_info["tel"]) 
                    set_data("I41", "K41", delivery_info["contact"]) 
                    
                    remark_lines = []
                    if option in ["チャーター便", "OLT"]:
                        remark_lines.append(f"{option}をお願いいたします。\n消防法に関わる危険物等の有無を確認し、問題ないよう手配をお願いします。")
                    if data.get("internal_memo"):
                        remark_lines.append(f"【社内メモ転記】\n{data.get('internal_memo')}")
                    
                    ws.merge_cells("C44:K49")
                    ws["C44"] = "\n\n".join(remark_lines)
                    ws["C44"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    ws.row_dimensions[44].height = 80

                    output = BytesIO()
                    wb.save(output)
                    
                    st.subheader("📁 Excelファイルのダウンロード")
                    st.download_button(
                        label="完成したExcelをダウンロード",
                        data=output.getvalue(),
                        file_name=f"集荷依頼書_{data.get('bl_no', 'unknown')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

# --- 画面切り替え：顧客連絡用テキスト作成 モード ---
elif mode == "顧客連絡用テキスト作成":
    st.header("📋 顧客連絡用テキスト作成")
    st.write("アップロードされたスクショ画像から、自動的に顧客案内の文面を作成します。")
    
    # ★追加：搬入先CFSのプルダウン選択
    selected_cfs = st.selectbox("搬入先CFSを選択してください", list(CFS_MASTER.keys()))
    
    if uploaded_file and api_key:
        if st.button("連絡用テキストを生成する"):
            with st.spinner("最新のGemini AIと通信してデータを抽出しています...（約10秒）"):
                data = call_gemini_api(api_key, uploaded_file)
                if data:
                    st.success("テキストの生成が完了しました！")
                    
                    # ★追加：港名から「国名」を省く処理
                    receipt_port = format_port_name(data.get('place_of_receipt', ''))
                    loading_port = format_port_name(data.get('place_of_loading', ''))
                    delivery_port = format_port_name(data.get('place_of_delivery', ''))
                    
                    # ★追加：選択されたCFSのテキストを取得
                    cfs_text = CFS_MASTER[selected_cfs]
                    
                    email_text = f"""下記のように承りました。
BOOKING NO. {data.get('bl_no', '')}
VESSEL NAME {data.get('vessel', '')} {data.get('voy', '')}
{receipt_port} CFS CUT : {data.get('cfs_cut', '')}
{loading_port} ETD : {data.get('etd', '')}
{delivery_port} ETA : {data.get('eta', '')}

{cfs_text}"""

                    st.code(email_text, language="text")
                    
                    with st.expander("抽出された生データ (JSON) を確認する"):
                        st.json(data)